#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import os
from openpyxl.styles import Border, Side, Font, colors


def get_headers(columns):
    alphabet = [
        "A",
        "B",
        "C",
        "D",
        "E",
        "F",
        "G",
        "H",
        "I",
        "J",
        "K",
        "L",
        "M",
        "N",
        "O",
        "P",
        "Q",
        "R",
        "S",
        "T",
        "U",
        "V",
        "W",
        "X",
        "Y",
        "Z",
    ]
    excel_cols = []

    if columns <= len(alphabet):
        excel_cols += alphabet[:columns]

    elif columns > len(alphabet):
        excel_cols += alphabet

        # AA, AB, AC..ZZ
        for letter in range(len(alphabet)):
            for second_letter in range(len(alphabet)):
                if len(excel_cols) >= columns:
                    return str(excel_cols[0]), str(excel_cols[-1])

                excel_cols.append(alphabet[letter] + alphabet[second_letter])

    return str(excel_cols[0]), str(excel_cols[-1])


def modify_sheet_title(wb, sheetname):
    ws = wb["Sheet"]
    ws.title = sheetname

    return ws


def get_columns(table_header, columns, force_columns, quick_create):
    if force_columns or quick_create:
        cols = columns
    else:
        try:
            cols = len(table_header)
        except TypeError:
            cols = columns
    return cols


def format_sheet(ws, rows, cols, accent_type):

    thin = Side(border_style="thin", color="000000")

    for row in ws.iter_rows(min_row=0, min_col=0, max_row=rows, max_col=cols):
        for cell in row:

            if cell.row is 1:
                cell.style = f"60 % - {accent_type}"
            else:
                cell.style = f"20 % - {accent_type}"

            cell.border = Border(left=thin, top=thin, right=thin, bottom=thin)
            cell.font = Font(color=colors.BLACK)

    return ws


def create_table_header(ws, cols, table_header):
    first_col, last_col = get_headers(cols)  # column numbers
    for row in ws[f"{first_col}1:{last_col}1"]:
        for i, cell in enumerate(row):
            if i >= len(table_header):
                return ws
            cell.value = table_header[i]
            cell.font = Font(color=colors.BLACK)

    return ws


# Force_columns, incase user wants to input some column names and leave formatted table space for more
def generate_xlsx(
    table_header,
    filename,
    filepath=None,
    sheetname="Sheet1",
    rows=4,
    columns=5,
    accent_type="Accent1",
    force_columns=False,
    quick_create=False,
):

    if filename:
        if "xlsx" not in filename:
            filename += ".xlsx"

    if not filepath:
        filepath = os.path.join(os.getcwd())

    if type(table_header) == str:
        table_header = table_header.split(",")
        table_header = list(map(lambda x: x.strip(), table_header))

    wb = openpyxl.Workbook()
    ws = modify_sheet_title(wb, sheetname)
    cols = get_columns(table_header, columns, force_columns, quick_create)
    ws = format_sheet(ws, rows, cols, accent_type)

    if quick_create:
        wb.save(f"{filepath}/{filename}")
        return

    if table_header:
        ws = create_table_header(ws, cols, table_header)

    wb.save(f"{filepath}/{filename}")


if __name__ == "__main__":
    print(
        "This file is NOT meant to be run directly. Please Use the CLI tool or the GUI tool, Thanks =)"
    )
